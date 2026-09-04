import csv
import json
import http.client
import threading
import tempfile
import unittest
from pathlib import Path
from unittest.mock import Mock, patch

import i18n

from functional_core import (
    ApiKeyError,
    Execution,
    LocalStore,
    ValidationError,
    calculate_carbon,
    classify_cpu_tier,
    calculate_cost,
    calculate_energy,
    calculate_execution,
    calculate_water,
    budget_percentage,
    component_percentages,
    compare_models,
    convert_clp,
    decrypt_api_key,
    encrypt_api_key,
    estimate_cloud,
    export_records,
    fetch_exchange_rates,
    format_carbon,
    green_score,
    import_records,
    forecast_budget,
    mask_api_key,
    rightsizing,
    semaphore_level,
    sanitize_markdown,
    render_markdown,
    validate_api_key_format,
    validate_password,
    validate_thresholds,
    hash_password,
)


class FunctionalCoreTests(unittest.TestCase):
    def test_calculation_pipeline(self):
        self.assertEqual(calculate_cost(12, 2, "USD"), 24.0)
        self.assertEqual(calculate_energy(1000, 2, 1.5), 3.0)
        self.assertEqual(calculate_carbon(1000, 2, 1, 400), 800.0)
        self.assertEqual(calculate_carbon(1000, 2, 1, 400, 1, 1000), 1400.0)
        self.assertEqual(calculate_water(3, 2, 3, immersion=True), 0.0)
        self.assertEqual(format_carbon(10001), "10.00 kgCO2eq")

    @patch("functional_core.requests.get")
    def test_exchange_rates_from_api_and_inverse_conversion(self, get):
        response = Mock()
        response.json.return_value = {
            "result": "success",
            "rates": {currency: index + 1 for index, currency in enumerate(("USD", "EUR", "BRL", "PEN", "ARS", "CNY", "GBP", "JPY", "CAD", "CHF"))},
        }
        get.return_value = response
        with tempfile.TemporaryDirectory() as directory:
            rates = fetch_exchange_rates(Path(directory) / "rates.json")
        self.assertEqual(rates["CLP"], 1.0)
        self.assertEqual(set(rates), {"CLP", "USD", "EUR", "BRL", "PEN", "ARS", "CNY", "GBP", "JPY", "CAD", "CHF"})
        self.assertEqual(convert_clp(100, "USD", rates), (100.0, 1.0))
        get.assert_called_once()

    @patch("functional_core.requests.get", side_effect=ConnectionError("offline"))
    def test_exchange_rates_use_local_fallback(self, get):
        rates_data = {currency: 0.5 for currency in ("USD", "EUR", "BRL", "PEN", "ARS", "CNY", "GBP", "JPY", "CAD", "CHF")}
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "rates.json"
            path.write_text(json.dumps({"rates": rates_data}), encoding="utf-8")
            rates = fetch_exchange_rates(path)
        self.assertEqual(rates["USD"], 0.5)

    def test_xlsx_report_contains_structured_sheets(self):
        from openpyxl import load_workbook
        from export_handler import _create_xlsx_report

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            _create_xlsx_report(
                "economia",
                {
                    "kpis": [[15, 60, "Costo", "100", "USD ($)", "cyan_500"]],
                    "details": [["GPU", "48%", "emerald_500"]],
                    "logs": [["OK", "red_500"]],
                    "exported_by": "Test",
                    "progress": 64,
                },
                path,
            )
            workbook = load_workbook(path)
            try:
                self.assertEqual(set(workbook.sheetnames), {"KPIs", "Detalles", "Actividad", "Resumen"})
                self.assertEqual(workbook["KPIs"]["A1"].value, "Posición")
                self.assertEqual(workbook["KPIs"]["E2"].value, "USD ($)")
                self.assertNotIn("Color", [cell.value for cell in workbook["KPIs"][1]])
                self.assertNotIn("Color", [cell.value for cell in workbook["Detalles"][1]])
                self.assertNotIn("Color", [cell.value for cell in workbook["Actividad"][1]])
                self.assertEqual(workbook["Resumen"]["B4"].value, "Test")
                self.assertEqual(workbook["KPIs"]["A1"].fill.fgColor.rgb[-6:], "17324D")
                self.assertEqual(workbook["KPIs"]["D2"].fill.fgColor.rgb[-6:], "CFFAFE")
                self.assertEqual(workbook["Detalles"]["A2"].fill.fgColor.rgb[-6:], "D1FAE5")
                self.assertEqual(workbook["Actividad"]["A2"].fill.fgColor.rgb[-6:], "FEE2E2")
                self.assertEqual(workbook["KPIs"].auto_filter.ref, "A1:E2")
            finally:
                workbook.close()

    def test_invalid_business_values(self):
        with self.assertRaises(ValidationError):
            validate_thresholds(60, 50, 90)
        with self.assertRaises(ValidationError):
            validate_password("weak")
        with self.assertRaises(ValidationError):
            calculate_water(1, 1, 4)

    def test_green_score_and_semaphore(self):
        score, badge = green_score(10, 100, 10, 100)
        self.assertEqual((score, badge), (90.0, "A+"))
        self.assertEqual(semaphore_level(40, 50, 90, 100), "Verde")
        self.assertEqual(semaphore_level(70, 50, 90, 100), "Amarillo")
        self.assertEqual(semaphore_level(95, 50, 90, 100), "Rojo")

    def test_comparison_ties_and_markdown_limits(self):
        compared = compare_models([
            {"name": "A", "carbon": 10, "cost": 2},
            {"name": "B", "carbon": 10, "cost": 3},
            {"name": "C", "carbon": 20, "cost": 1},
        ])
        self.assertEqual([row["optimal"] for row in compared], [True, True, False])
        self.assertEqual(sanitize_markdown("<script>alert(1)</script>"), "&lt;script&gt;alert(1)&lt;/script&gt;")
        with self.assertRaises(ValidationError):
            sanitize_markdown("x" * 4, max_chars=3)

    def test_safe_markdown_and_component_percentages(self):
        html = render_markdown("# Titulo\n\n- **CPU**\n- [Sitio](https://example.com)\n<script>alert(1)</script>")
        self.assertIn("<h1>Titulo</h1>", html)
        self.assertIn("<strong>CPU</strong>", html)
        self.assertIn('href="https://example.com"', html)
        self.assertNotIn("<script>", html)
        percentages = component_percentages({"CPU": 1, "GPU": 2, "RAM": 1})
        self.assertEqual(sum(percentages.values()), 100.0)
        self.assertEqual(percentages["GPU"], 50.0)
        self.assertEqual(budget_percentage(125, 100), 100)
        self.assertIsNone(budget_percentage(125, 0))
        with self.assertRaises(ValidationError):
            budget_percentage(-1, 100)

    def test_api_key_encryption_roundtrip_and_validation(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            key_path = Path(tmp_dir) / "secrets" / "financial_api.key"
            token = encrypt_api_key("ABCDEFGHIJKLMNOP1234", key_path)
            self.assertNotIn("ABCDEFGHIJKLMNOP1234", token)
            self.assertEqual(decrypt_api_key(token, key_path), "ABCDEFGHIJKLMNOP1234")
            self.assertEqual(mask_api_key("ABCDEFGHIJKLMNOP1234"), "****1234")
            with self.assertRaises(ApiKeyError):
                validate_api_key_format("short")
            with self.assertRaises(ApiKeyError):
                decrypt_api_key("not-a-real-token", key_path)

    def test_rightsizing_and_budget_forecast(self):
        recommendation = rightsizing(200, [
            {"name": "same", "tdp_watts": 190},
            {"name": "efficient", "tdp_watts": 150},
        ])
        self.assertEqual(recommendation["candidate"]["name"], "efficient")
        self.assertEqual(recommendation["saving_percent"], 25.0)
        self.assertEqual(forecast_budget(500, 15, 800, 30), 1000.0)
        with self.assertRaises(ValidationError):
            forecast_budget(500, 0, 800, 30)

    def test_cpu_tier_classification_and_rightsizing_filter(self):
        self.assertEqual(classify_cpu_tier("Atom C2750"), 0)
        self.assertEqual(classify_cpu_tier("Core i5-10300H"), 2)
        self.assertEqual(classify_cpu_tier("Ryzen 7 5800X"), 3)
        self.assertEqual(classify_cpu_tier("Core i9-14900K"), 4)

        current_tier = classify_cpu_tier("Core i5-10300H")
        candidates = [
            {"name": "Atom C2750", "tdp_watts": 20, "performance_score": classify_cpu_tier("Atom C2750")},
            {"name": "Core i5 Efficient", "tdp_watts": 30, "performance_score": classify_cpu_tier("Core i5-9400F")},
        ]
        recommendation = rightsizing(45, candidates, current_performance=current_tier)
        self.assertEqual(recommendation["candidate"]["name"], "Core i5 Efficient")

    def test_new_ui_strings_are_bilingual(self):
        translations = {
            "Comparativa de modelos": "Model comparison",
            "Empate técnico": "Technical tie",
            "Sugerir hardware eficiente": "Suggest efficient hardware",
            "No hay ejecuciones registradas.": "No executions recorded.",
            "Usuario bloqueado": "User locked",
        }
        for spanish, english in translations.items():
            self.assertEqual(i18n.t(spanish, "en"), english)
            self.assertEqual(i18n.t(english, "es"), spanish)

    def test_cloud_traceability_and_execution(self):
        estimate = estimate_cloud({"name": "small", "cost_per_hour_usd": 2, "watts": 500}, 3, 400)
        self.assertEqual(estimate["cost_usd"], 6.0)
        self.assertEqual(estimate["carbon_gco2eq"], 600.0)
        self.assertEqual(estimate["inputs"]["watts"], 500.0)
        execution, badge = calculate_execution(
            model_id=1, hourly_cost=2, hours=3, currency="USD", tdp_watts=500,
            pue=1.2, grid_factor=400, wue=1, wsi=1, cost_limit=100,
            carbon_limit=1000,
        )
        self.assertEqual(execution.kwh, 1.8)
        self.assertEqual(execution.carbon, 720.0)
        self.assertEqual(badge, "C")

    def test_clear_and_delete_project_are_isolated(self):
        with tempfile.TemporaryDirectory() as directory:
            store = LocalStore(Path(directory) / "projects.sqlite3")
            first_project = store.add_project("Primero")
            second_project = store.add_project("Segundo")
            first_model = store.add_model(first_project, "Modelo A")
            second_model = store.add_model(second_project, "Modelo B")
            for model_id in (first_model, second_model):
                store.connection.execute(
                    """INSERT INTO executions(
                           model_id, timestamp, cost, carbon, kwh, water, duration_ms, semaphore
                       ) VALUES (?, '2026-01-01 00:00:00', 1, 2, 3, 4, 5, 'Verde')""",
                    (model_id,),
                )
            store.connection.commit()

            store.clear_project(first_project)
            self.assertEqual(len(store.list_models(first_project)), 0)
            self.assertEqual(len(store.list_history(project_id=first_project)), 0)
            self.assertEqual(len(store.list_models(second_project)), 1)
            self.assertEqual(len(store.list_history(project_id=second_project)), 1)
            self.assertIsNotNone(store.connection.execute(
                "SELECT id FROM projects WHERE id = ?", (first_project,)
            ).fetchone())

            store.delete_project(first_project)
            self.assertIsNone(store.connection.execute(
                "SELECT id FROM projects WHERE id = ?", (first_project,)
            ).fetchone())
            self.assertEqual(len(store.list_projects()), 1)
            with self.assertRaises(ValidationError):
                store.delete_project(first_project)
            store.close()

    def test_secure_authentication_and_lockout(self):
        with tempfile.TemporaryDirectory() as directory:
            store = LocalStore(Path(directory) / "test.sqlite3")
            store.add_user("admin", "ClaveSegura1@", "admin")
            self.assertEqual(store.authenticate("admin", "ClaveSegura1@")["role"], "admin")
            for _ in range(5):
                self.assertIsNone(store.authenticate("admin", "incorrecta"))
            self.assertIsNone(store.authenticate("admin", "ClaveSegura1@"))
            row = store.connection.execute(
                "SELECT failed_attempts, is_locked FROM users WHERE username = ?", ("admin",)
            ).fetchone()
            self.assertEqual((row["failed_attempts"], row["is_locked"]), (5, 1))
            store.close()

    def test_server_authentication_uses_shared_store_and_lockout(self):
        import server

        with tempfile.TemporaryDirectory() as directory:
            database_path = Path(directory) / "server.sqlite3"
            store = LocalStore(database_path)
            store.add_user("remote", "ClaveSegura1@")
            store.close()

            def open_store():
                return LocalStore(database_path)

            with patch("server.get_store", side_effect=open_store):
                user, error, status = server.authenticate_request("remote", "ClaveSegura1@")
                self.assertEqual((user["username"], error, status), ("remote", None, 200))
                for _ in range(5):
                    user, error, status = server.authenticate_request("remote", "incorrecta")
                self.assertEqual((user, error, status), (None, "Usuario bloqueado", 423))

    def test_server_http_login_token_and_hardware_access(self):
        import server

        with tempfile.TemporaryDirectory() as directory:
            database_path = Path(directory) / "server_http.sqlite3"
            store = LocalStore(database_path)
            password_hash = hash_password("ClaveSegura1@")
            store.add_hashed_user("remote", password_hash)
            store.close()

            def open_store():
                return LocalStore(database_path)

            httpd = server.socketserver.ThreadingTCPServer(("127.0.0.1", 0), server.SimpleHandler)
            thread = threading.Thread(target=httpd.serve_forever, daemon=True)
            thread.start()
            try:
                with patch("server.get_store", side_effect=open_store), patch(
                    "server.load_config",
                    return_value={"users": [{"username": "remote", "password_hash": password_hash}]},
                ), patch("server.get_hardware_info", return_value={"cpu": "test"}):
                    connection = http.client.HTTPConnection("127.0.0.1", httpd.server_address[1])
                    body = json.dumps({"username": "remote", "password": "ClaveSegura1@"})
                    connection.request("POST", "/login", body, {"Content-Type": "application/json"})
                    response = connection.getresponse()
                    login_data = json.loads(response.read().decode("utf-8"))
                    self.assertEqual(response.status, 200)
                    self.assertTrue(login_data["token"])

                    connection.request("GET", "/hardware")
                    self.assertEqual(connection.getresponse().status, 401)
                    connection.request("GET", "/hardware", headers={"Authorization": f"Bearer {login_data['token']}"})
                    hardware_response = connection.getresponse()
                    self.assertEqual((hardware_response.status, json.loads(hardware_response.read())["cpu"]), (200, "test"))
                    connection.close()
            finally:
                httpd.shutdown()
                httpd.server_close()
                thread.join(timeout=2)

    def test_admin_can_view_and_unlock_accounts(self):
        with tempfile.TemporaryDirectory() as directory:
            store = LocalStore(Path(directory) / "users.sqlite3")
            store.add_user("maxine", "ClaveSegura1@")
            for _ in range(5):
                store.authenticate("maxine", "incorrecta")
            status = store.list_user_status()[0]
            self.assertEqual((status["failed_attempts"], status["is_locked"]), (5, 1))
            with self.assertRaises(PermissionError):
                store.unlock_user("maxine", "Usuario")
            store.unlock_user("maxine", "Administrador")
            status = store.list_user_status()[0]
            self.assertEqual((status["failed_attempts"], status["is_locked"]), (0, 0))
            store.close()

    def test_json_round_trip_and_corruption(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "records.json"
            records = [{"name": "modelo", "is_active": True}]
            export_records(records, path)
            self.assertEqual(import_records(path), records)
            path.write_text("{bad", encoding="utf-8")
            with self.assertRaises(ValidationError):
                import_records(path)

    def test_project_lifecycle_and_reassignment(self):
        with tempfile.TemporaryDirectory() as directory:
            store = LocalStore(Path(directory) / "projects.sqlite3")
            source = store.add_project("Proyecto A")
            target = store.add_project("Proyecto B")
            model = store.add_model(source, "Modelo QA", "# Descripcion")
            store.add_execution(Execution(model, "2026-08-24 12:00:00", 10, 20, 2, 3, 900, "Verde"))
            self.assertEqual(store.project_totals(source)["cost"], 10.0)
            store.reassign_model(model, target)
            self.assertEqual(store.project_totals(source)["cost"], 0.0)
            self.assertEqual(store.project_totals(target)["cost"], 10.0)
            store.archive_project(target)
            with self.assertRaises(ValidationError):
                store.add_model(target, "Bloqueado")
            backup = Path(directory) / "backup.sqlite3"
            store.backup(backup)
            self.assertTrue(backup.exists())
            backup_store = LocalStore(backup)
            self.assertEqual(backup_store.project_totals(target)["cost"], 10.0)
            self.assertEqual(len(backup_store.list_history()), 1)
            backup_store.close()
            store.close()

    def test_empty_history_is_explicit(self):
        with tempfile.TemporaryDirectory() as directory:
            store = LocalStore(Path(directory) / "history.sqlite3")
            self.assertEqual(store.list_history(), [])
            store.close()

    def test_hardware_catalog_has_real_ram_options(self):
        catalog_path = Path(__file__).parent / "data" / "hardware.csv"
        with catalog_path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        ram_rows = [row for row in rows if row["Tipo_Componente"] == "RAM"]
        self.assertGreaterEqual(len(ram_rows), 5)
        self.assertTrue(all(row["Capacidad_GB"] for row in ram_rows))
        self.assertFalse(any("prueba" in row["Modelo"].lower() for row in rows))


if __name__ == "__main__":
    unittest.main()
