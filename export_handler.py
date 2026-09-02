import os
import sys
from pathlib import Path
from PySide6.QtCore import QObject, QThread, Qt, QStandardPaths, Signal, Slot
from PySide6.QtWidgets import QApplication, QMessageBox, QFileDialog

import i18n
from i18n import t
from app_paths import resource_path, writable_path

# Add export handler to path
sys.path.append(resource_path("export handler"))

try:
    import eco
    import economia
    import inicio
    _EXPORT_IMPORT_ERROR = None
except Exception as exc:
    eco = None
    economia = None
    inicio = None
    _EXPORT_IMPORT_ERROR = exc

# Mantiene vivas las referencias a hilos/workers en curso (si no, Python los recolecta a mitad de la tarea).
_active_exports = []


def _create_xlsx_report(report_type, data, file_path):
    """Write a compact, styled workbook that mirrors the PDF report hierarchy."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    colors = {
        "navy": "17324D",
        "cyan": "06B6D4",
        "cyan_light": "CFFAFE",
        "green": "059669",
        "green_light": "D1FAE5",
        "amber": "D97706",
        "amber_light": "FEF3C7",
        "red": "DC2626",
        "gray": "6B7280",
        "gray_light": "F3F4F6",
        "border": "D9E1E8",
        "white": "FFFFFF",
    }
    thin_border = Border(bottom=Side(style="thin", color=colors["border"]))
    header_fill = PatternFill("solid", fgColor=colors["navy"])
    accent_fill = PatternFill("solid", fgColor=colors["cyan_light"])
    row_fills = {
        "cyan_500": PatternFill("solid", fgColor=colors["cyan_light"]),
        "cyan_600": PatternFill("solid", fgColor=colors["cyan_light"]),
        "emerald_500": PatternFill("solid", fgColor=colors["green_light"]),
        "emerald_600": PatternFill("solid", fgColor=colors["green_light"]),
        "amber_500": PatternFill("solid", fgColor=colors["amber_light"]),
        "red_500": PatternFill("solid", fgColor="FEE2E2"),
        "gray_800": PatternFill("solid", fgColor=colors["gray_light"]),
        "gray_500": PatternFill("solid", fgColor=colors["gray_light"]),
        "logo_orange": PatternFill("solid", fgColor="FFEDD5"),
    }
    row_fonts = {
        "cyan_500": colors["cyan"],
        "cyan_600": colors["cyan"],
        "emerald_500": colors["green"],
        "emerald_600": colors["green"],
        "amber_500": colors["amber"],
        "red_500": colors["red"],
        "gray_800": colors["navy"],
        "gray_500": colors["gray"],
        "logo_orange": "C2410C",
    }

    def style_sheet(sheet, headers, rows, widths):
        sheet.sheet_view.showGridLines = False
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(name="Aptos", bold=True, color=colors["white"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 24
        for row in rows:
            sheet.append(list(row))
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row in sheet.iter_rows(min_row=2):
            if row:
                row[0].font = Font(name="Aptos", color=colors["navy"])
        return sheet

    def style_data_rows(sheet, source_rows, color_index, emphasis_columns=None):
        emphasis_columns = emphasis_columns or range(1, sheet.max_column + 1)
        for row_number, source_row in enumerate(source_rows, start=2):
            key = str(source_row[color_index]) if len(source_row) > color_index else ""
            fill = row_fills.get(key)
            font_color = row_fonts.get(key)
            if fill:
                for cell in sheet[row_number]:
                    cell.fill = fill
            if font_color:
                for column in emphasis_columns:
                    cell = sheet.cell(row_number, column)
                    cell.font = Font(name="Aptos", bold=column in {1, 3, 4}, color=font_color)

    workbook = Workbook()
    source_kpis = data.get("kpis", [])
    kpi_rows = [[row[0], row[1], row[2], row[3], row[4]] for row in source_kpis]
    kpis = style_sheet(workbook.active, ["Posición", "Ancho", "Indicador", "Valor", "Unidad"], kpi_rows, [12, 12, 28, 20, 16])
    kpis.title = "KPIs"
    style_data_rows(kpis, source_kpis, 5, emphasis_columns=(3, 4, 5))

    source_details = data.get("details", [])
    detail_rows = [[row[0], row[1]] for row in source_details]
    details = style_sheet(workbook.create_sheet("Detalles"), ["Detalle", "Valor"], detail_rows, [32, 24])
    style_data_rows(details, source_details, 2, emphasis_columns=(1, 2))
    source_logs = data.get("logs", [])
    log_rows = [[row[0]] for row in source_logs]
    logs = style_sheet(workbook.create_sheet("Actividad"), ["Mensaje"], log_rows, [76])
    style_data_rows(logs, source_logs, 1, emphasis_columns=(1,))

    summary = workbook.create_sheet("Resumen", 0)
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:D1")
    summary["A1"] = "Semáforo IA"
    summary["A1"].font = Font(name="Aptos Display", bold=True, size=20, color=colors["navy"])
    summary["A1"].fill = header_fill
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 34
    summary.append(["Resumen del informe", "Valor", "", ""])
    summary.append(["Tipo de reporte", report_type, "", ""])
    summary.append(["Exportado por", data.get("exported_by", ""), "", ""])
    summary.append(["Progreso", data.get("progress", ""), "%", ""])
    summary.append(["KPIs incluidos", len(data.get("kpis", [])), "", ""])
    summary.append(["Detalles incluidos", len(data.get("details", [])), "", ""])
    summary.append(["Registros de actividad", len(data.get("logs", [])), "", ""])
    for cell in summary[2]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color=colors["white"])
    for row in range(3, summary.max_row + 1):
        summary.cell(row, 1).font = Font(bold=True, color=colors["navy"])
        summary.cell(row, 1).fill = accent_fill
        for cell in summary[row]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    summary["B5"].number_format = "0%"
    summary["B5"] = float(data.get("progress", 0) or 0) / 100
    summary.conditional_formatting.add("B5", ColorScaleRule(start_type="min", start_color="FEE2E2", mid_type="percentile", mid_value=50, mid_color="FEF3C7", end_type="max", end_color="D1FAE5"))
    summary.freeze_panes = "A3"
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 32
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 4
    workbook.save(file_path)


def _send_os_notification():
    """Send a best-effort native notification when enabled in application config."""
    try:
        config_path = writable_path("config.json")
        if os.path.isfile(config_path):
            import json
            with open(config_path, "r", encoding="utf-8") as handle:
                if not json.load(handle).get("notifications_os", True):
                    return
        from plyer import notification
        notification.notify(
            title="Semaforo IA",
            message="Evaluacion completada",
            app_name="Semaforo IA",
            timeout=5,
        )
    except Exception:
        pass

def _pick_export_target(parent_widget, report_type, export_format, lang=None):
    if export_format == "json":
        ext = ".json"
        file_filter = t("Archivo JSON (*.json);;Todos los archivos (*)", lang)
    elif export_format == "xlsx":
        ext = ".xlsx"
        file_filter = t("Archivo Excel (*.xlsx);;Todos los archivos (*)", lang)
    else:
        ext = ".pdf"
        file_filter = t("Archivo PDF (*.pdf);;Todos los archivos (*)", lang)

    default_names = {
        "eco": f"informe_semaforo_ia{ext}",
        "economia": f"informe_economia_semaforo_ia{ext}",
        "inicio": f"informe_inicio_semaforo_ia{ext}",
        "proyecto": f"informe_proyecto_semaforo_ia{ext}",
    }
    default_name = default_names.get(report_type, f"informe_semaforo_ia{ext}")

    documents_dir = QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation) or os.path.expanduser("~")

    selected_path, _ = QFileDialog.getSaveFileName(
        parent_widget,
        t("Guardar Informe", lang),
        os.path.join(documents_dir, default_name),
        file_filter,
    )

    if not selected_path:
        return ""

    if export_format == "json" and not selected_path.lower().endswith(".json"):
        selected_path += ".json"
    elif export_format == "xlsx" and not selected_path.lower().endswith(".xlsx"):
        selected_path += ".xlsx"
    elif export_format in {"pdf", "both"} and not selected_path.lower().endswith(".pdf"):
        selected_path += ".pdf"

    return selected_path


def _apply_report_data(report_type, data, file_path, export_format, lang):
    """Actualiza el modulo de reporte y genera el PDF/JSON. Pensada para correr en un hilo aparte."""
    if export_format == "xlsx":
        _create_xlsx_report(report_type, data, file_path)
        return

    if _EXPORT_IMPORT_ERROR is not None:
        raise RuntimeError(f"No se pudieron cargar los módulos de exportación: {_EXPORT_IMPORT_ERROR}") from _EXPORT_IMPORT_ERROR

    if report_type in ('eco', 'proyecto'):
        # Update eco module data ('proyecto' reuses the same generic KPI/details/logs template)
        if "kpis" in data:
            eco.REPORT["kpis"] = data["kpis"]
        if "details" in data:
            eco.REPORT["details"] = data["details"]
        if "logs" in data:
            eco.REPORT["logs"] = data["logs"]
        if "progress" in data:
            eco.REPORT["progress"] = data["progress"]
        if "exported_by" in data:
            eco.SHARED["exported_by"] = data["exported_by"]

        if "chart_values" in data:
            eco.REPORT["chart_values"] = data["chart_values"]
        if "chart_labels" in data:
            eco.REPORT["chart_labels"] = data["chart_labels"]
        eco.create_pdf_report(filename=file_path, export_format=export_format, lang=lang)

    elif report_type == 'economia':
        if "kpis" in data:
            economia.REPORT["kpis"] = data["kpis"]
        if "details" in data:
            economia.REPORT["details"] = data["details"]
        if "logs" in data:
            economia.REPORT["logs"] = data["logs"]
        if "progress" in data:
            economia.REPORT["progress"] = data["progress"]
        if "exported_by" in data:
            economia.SHARED["exported_by"] = data["exported_by"]
        if "chart_values" in data:
            economia.REPORT["chart_values"] = data["chart_values"]
        if "chart_labels" in data:
            economia.REPORT["chart_labels"] = data["chart_labels"]

        economia.create_pdf_report(filename=file_path, export_format=export_format, lang=lang)

    elif report_type == 'inicio':
        if "kpis" in data:
            inicio.REPORT["kpis"] = data["kpis"]
        if "details" in data:
            inicio.REPORT["details"] = data["details"]
        if "logs" in data:
            inicio.REPORT["logs"] = data["logs"]
        if "progress" in data:
            inicio.REPORT["progress"] = data["progress"]
        if "badge" in data:
            inicio.REPORT["badge"] = data["badge"]
        if "accent_color" in data:
            inicio.REPORT["accent_color"] = data["accent_color"]
        if "accent_color_dark" in data:
            inicio.REPORT["accent_color_dark"] = data["accent_color_dark"]
        if "accent_color_light" in data:
            inicio.REPORT["accent_color_light"] = data["accent_color_light"]
        if "exported_by" in data:
            inicio.SHARED["exported_by"] = data["exported_by"]
        if "chart_values" in data:
            inicio.REPORT["chart_values"] = data["chart_values"]
        if "chart_labels" in data:
            inicio.REPORT["chart_labels"] = data["chart_labels"]
        if "chart_colors" in data:
            inicio.REPORT["chart_colors"] = data["chart_colors"]

        inicio.create_pdf_report(filename=file_path, export_format=export_format, lang=lang)


class _ExportWorker(QObject):
    finished = Signal(bool, str)

    def __init__(self, report_type, data, file_path, export_format, lang):
        super().__init__()
        self._report_type = report_type
        self._data = data
        self._file_path = file_path
        self._export_format = export_format
        self._lang = lang

    def run(self):
        try:
            _apply_report_data(self._report_type, self._data, self._file_path, self._export_format, self._lang)
        except Exception as exc:
            self.finished.emit(False, str(exc))
        else:
            self.finished.emit(True, "")


class _ExportController(QObject):
    """Vive en el hilo principal: conectar una senal cruzando hilos a un slot de un
    QObject (en vez de a una funcion Python suelta) es lo que permite a Qt detectar
    la diferencia de hilo y encolar la llamada correctamente en el hilo de la UI."""

    def __init__(self, parent_widget, thread, worker, trigger_widget, lang, export_entry):
        super().__init__()
        self.parent_widget = parent_widget
        self.thread = thread
        self.worker = worker
        self.trigger_widget = trigger_widget
        self.lang = lang
        self.export_entry = export_entry

    @Slot(bool, str)
    def on_finished(self, success, error_message):
        QApplication.restoreOverrideCursor()
        if self.trigger_widget is not None:
            self.trigger_widget.setEnabled(True)

        if success:
            QMessageBox.information(
                self.parent_widget,
                t("Éxito", self.lang),
                t("El reporte ha sido generado y guardado exitosamente.", self.lang)
            )
        else:
            QMessageBox.critical(
                self.parent_widget,
                t("Error", self.lang),
                t("Ocurrió un error al generar el archivo:\n{error}", self.lang).format(error=error_message)
            )

        if success:
            _send_os_notification()

        self.thread.quit()
        self.thread.wait()
        if self.export_entry in _active_exports:
            _active_exports.remove(self.export_entry)


def generate_and_save_report(parent_widget, report_type, data, export_format="pdf", lang=None, trigger_widget=None):
    """
    report_type: 'eco', 'economia' or 'inicio'
    data: dictionary containing dynamic info to update the report.
    lang: idioma a usar para el PDF/JSON (por defecto, el idioma actual de la UI).
    trigger_widget: boton que dispara la exportacion; se deshabilita mientras se genera
    el reporte en segundo plano para evitar que la UI se congele o se dispare dos veces.
    """
    if export_format not in {"pdf", "json", "xlsx", "both"}:
        raise ValueError("export_format debe ser 'pdf', 'json', 'xlsx' o 'both'")

    lang = lang or i18n.get_language()

    file_path = _pick_export_target(parent_widget, report_type, export_format, lang=lang)

    if not file_path:
        return

    if trigger_widget is not None:
        trigger_widget.setEnabled(False)
    QApplication.setOverrideCursor(Qt.WaitCursor)

    thread = QThread(parent_widget)
    worker = _ExportWorker(report_type, data, file_path, export_format, lang)
    worker.moveToThread(thread)
    thread.started.connect(worker.run)

    export_entry = {"thread": thread, "worker": worker}
    _active_exports.append(export_entry)

    # El controller se crea (y se queda) en el hilo principal: eso es lo que hace que
    # la conexion de abajo se encole automaticamente en vez de ejecutarse en el hilo worker.
    controller = _ExportController(parent_widget, thread, worker, trigger_widget, lang, export_entry)
    export_entry["controller"] = controller
    worker.finished.connect(controller.on_finished)
    thread.finished.connect(thread.deleteLater)
    thread.start()
