import json
from datetime import datetime
from pathlib import Path

from config_loader import load_config

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None


def pdf_text(value):
    return str(value).encode("latin-1", "replace").decode("latin-1")


def _fmt(value, decimals=2):
    try:
        return f"{float(value):.{decimals}f}"
    except (TypeError, ValueError):
        return "0.00"


def _duration(value):
    try:
        seconds = max(0, float(value)) / 1000
    except (TypeError, ValueError):
        seconds = 0
    return f"{seconds:.2f} s"


def _energy(value):
    try:
        kwh = float(value)
    except (TypeError, ValueError):
        kwh = 0
    return f"{kwh * 1000:.2f} Wh" if abs(kwh) < 1 else f"{kwh:.2f} kWh"


def export_json_report(filename, data):
    payload = {
        "report_type": "proyecto",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "exported_by": data.get("exported_by", ""),
        "scope": data.get("scope", {}),
        "totals": data.get("totals", {}),
        "projects": data.get("projects", []),
        "executions": data.get("executions", []),
    }
    target = Path(filename).with_suffix(".json")
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(target)


def create_xlsx_report(filename, data):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(name="Aptos", bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="D9E1E8"))

    def style_header(sheet):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False

    summary = workbook.active
    summary.title = "Resumen"
    summary.append(["Campo", "Valor"])
    scope = data.get("scope", {})
    totals = data.get("totals", {})
    summary_rows = [
        ["Tipo de reporte", "proyecto"],
        ["Alcance", scope.get("project_name", "Todos los proyectos")],
        ["ID del proyecto", scope.get("project_id", "")],
        ["Exportado por", data.get("exported_by", "")],
        ["Ejecuciones", totals.get("execution_count", 0)],
        ["Costo (USD)", totals.get("cost", 0)],
        ["Carbono (gCO2eq)", totals.get("carbon", 0)],
        ["Energia (kWh)", totals.get("kwh", 0)],
        ["Agua (L)", totals.get("water", 0)],
    ]
    for row in summary_rows:
        summary.append(row)
    style_header(summary)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 36

    executions = workbook.create_sheet("Ejecuciones")
    executions.append(["ID", "Fecha", "Modelo", "Semaforo", "Costo USD", "Carbono gCO2eq", "Energia kWh", "Agua L", "Duracion ms"])
    for item in data.get("executions", []):
        executions.append([
            item.get("id", ""), item.get("timestamp", ""), item.get("model_name", ""),
            item.get("semaphore", ""), item.get("cost", 0), item.get("carbon", 0),
            item.get("kwh", 0), item.get("water", 0), item.get("duration_ms", 0),
        ])
    style_header(executions)
    for column, width in zip("ABCDEFGHI", [10, 22, 24, 14, 14, 18, 14, 12, 14]):
        executions.column_dimensions[column].width = width

    projects = workbook.create_sheet("Proyectos")
    projects.append(["ID", "Proyecto", "Costo USD", "Carbono gCO2eq", "Energia kWh", "Agua L", "Ejecuciones"])
    for item in data.get("projects", []):
        projects.append([item.get("id", ""), item.get("name", ""), item.get("cost", 0), item.get("carbon", 0), item.get("kwh", 0), item.get("water", 0), item.get("execution_count", 0)])
    style_header(projects)
    for column, width in zip("ABCDEFG", [10, 28, 14, 18, 14, 12, 14]):
        projects.column_dimensions[column].width = width

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    workbook.save(filename)


def create_pdf_report(filename, data, lang=None):
    if FPDF is None:
        raise RuntimeError("Falta instalar fpdf2 para exportar proyectos a PDF.")
    shared, report, colors = load_config("eco")
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_fill_color(247, 250, 252)
    pdf.rect(0, 0, 210, 297, "F")

    scope = data.get("scope", {})
    totals = data.get("totals", {})
    title = scope.get("project_name", "Todos los proyectos")
    pdf.set_text_color(23, 50, 77)
    pdf.set_font("helvetica", "B", 20)
    pdf.cell(0, 12, pdf_text("Reporte de proyecto"), ln=True)
    pdf.set_font("helvetica", "", 11)
    pdf.set_text_color(75, 85, 99)
    pdf.cell(0, 7, pdf_text(f"Proyecto: {title}"), ln=True)
    pdf.cell(0, 7, pdf_text(f"Exportado por: {data.get('exported_by', '')}"), ln=True)
    pdf.ln(5)

    kpis = [
        ("Costo total", f"{_fmt(totals.get('cost'))} USD"),
        ("Carbono total", f"{_fmt(totals.get('carbon'), 4)} gCO2eq"),
        ("Energia total", _energy(totals.get("kwh"))),
        ("Agua total", f"{_fmt(totals.get('water'), 4)} L"),
        ("Ejecuciones", str(totals.get("execution_count", 0))),
    ]
    pdf.set_font("helvetica", "B", 9)
    for index, (label, value) in enumerate(kpis):
        x = 15 + (index % 3) * 62
        y = pdf.get_y() + (index // 3) * 25
        pdf.set_fill_color(255, 255, 255)
        pdf.set_draw_color(210, 220, 228)
        pdf.rect(x, y, 56, 19, "DF")
        pdf.set_xy(x + 4, y + 3)
        pdf.set_text_color(107, 114, 128)
        pdf.cell(48, 4, pdf_text(label))
        pdf.set_xy(x + 4, y + 9)
        pdf.set_text_color(23, 50, 77)
        pdf.set_font("helvetica", "B", 11)
        pdf.cell(48, 5, pdf_text(value))
        pdf.set_font("helvetica", "B", 9)
    pdf.set_y(pdf.get_y() + 34)

    pdf.set_text_color(23, 50, 77)
    pdf.set_font("helvetica", "B", 13)
    pdf.cell(0, 8, pdf_text("Ejecuciones registradas"), ln=True)
    headers = ["Fecha", "Modelo", "Estado", "Costo", "Carbono", "Energia", "Duracion"]
    widths = [31, 37, 20, 20, 27, 25, 25]
    pdf.set_fill_color(23, 50, 77)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 7)
    for header, width in zip(headers, widths):
        pdf.cell(width, 7, pdf_text(header), border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_font("helvetica", "", 7)
    for item in data.get("executions", []):
        values = [
            item.get("timestamp", ""), item.get("model_name", ""), item.get("semaphore", ""),
            f"{_fmt(item.get('cost'))} USD", f"{_fmt(item.get('carbon'), 4)} g",
            _energy(item.get("kwh")), _duration(item.get("duration_ms")),
        ]
        pdf.set_text_color(31, 41, 55)
        for value, width in zip(values, widths):
            pdf.cell(width, 7, pdf_text(value), border=1)
        pdf.ln()
    if not data.get("executions"):
        pdf.cell(sum(widths), 8, pdf_text("No hay ejecuciones registradas."), border=1, align="C")
        pdf.ln()

    pdf.ln(7)
    pdf.set_font("helvetica", "I", 8)
    pdf.set_text_color(107, 114, 128)
    pdf.cell(0, 5, pdf_text("Generado por Semaforo IA. Valores calculados desde el proyecto seleccionado."), align="C")
    pdf.output(filename)
